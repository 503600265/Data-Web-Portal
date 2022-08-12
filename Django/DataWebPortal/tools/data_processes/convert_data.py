import pandas as pd
import openpyxl
from openpyxl.workbook import Workbook
import pyarrow
# import win32com.client as win32
import xlrd
from pathlib import Path
import json
import os

def convert_document(input_dir, output_dir, output_type):
    file_name, file_extension = os.path.splitext(input_dir)
    input_type = file_extension
    if input_type == ".csv":
        if output_type == 'xlsx':
            csv_to_xlsx(input_dir,output_dir)
        if output_type == 'parquet':
            csv_parquet(input_dir,output_dir)
    if input_type == ".parquet":
        if output_type == 'csv':
            parquet_csv(input_dir,output_dir)
    if input_type == ".xlsx":
        if output_type == 'parquet':
            xlsx_parquet(input_dir,output_dir)
    if input_type == ".txt":
        if output_type == 'csv':
            txt_csv(input_dir,output_dir)
        if output_type == 'xlsx':
            txt_xlsx(input_dir,output_dir)
    if input_type == ".xls":
        if output_type == 'xlsx':
            xls_xlsx(input_dir,output_dir)
    if input_type == ".json":
        if output_type == 'csv':
            json_csv(input_dir,output_dir)
        if output_type == 'xlsx':
            json_xlsx(input_dir,output_dir)



def csv_to_xlsx(input_path, output_path):
    read_file = pd.read_csv (r'{}'.format(input_path))
    read_file.to_excel (r'{}'.format(output_path), index = None, header=True)

def csv_to_xlsx(input_path, output_path):
    df = pd.read_csv(r'{}'.format(input_path))
    df.to_parquet(r'{}'.format(output_path))

def parquet_csv(input_path, output_path):
    df = pd.read_parquet(r'{}'.format(input_path))
    df.to_csv(r'{}'.format(output_path), index = None)

def xlsx_parquet(input_path, output_path):
    df = pd.read_excel(r'{}'.format(input_path))
    df.to_parquet(r'{}'.format(output_path))

def txt_csv(input_path, output_path):
    read_file = pd.read_csv (r'{}'.format(input_path))
    read_file.to_csv (r'{}'.format(output_path), index=None)

def txt_xlsx(input_path, output_path):
    read_file = pd.read_csv (r'{}'.format(input_path))
    read_file.to_excel (r'{}'.format(output_path), index = None, header=True)

def xls_xlsx(input_path,output_path):
#Windows
    # excel = win32.gencache.EnsureDispatch('Excel.Application')
    # wb = excel.Workbooks.Open(input_path)
    # wb.SaveAs(output_path, FileFormat = 51)    #FileFormat = 51 is for .xlsx extension
    # wb.Close()                               #FileFormat = 56 is for .xls extension
    # excel.Application.Quit()
#Linux    
    book_xls = xlrd.open_workbook(input_path)
    book_xlsx = Workbook()
    sheet_names = book_xls.sheet_names()
    for sheet_index, sheet_name in enumerate(sheet_names):
        sheet_xls = book_xls.sheet_by_name(sheet_name)
        if sheet_index == 0:
            sheet_xlsx = book_xlsx.active
            sheet_xlsx.title = sheet_name
        else:
            sheet_xlsx = book_xlsx.create_sheet(title=sheet_name)
        for row in range(0, sheet_xls.nrows):
            for col in range(0, sheet_xls.ncols):
                sheet_xlsx.cell(row = row+1 , column = col+1).value = sheet_xls.cell_value(row, col)
    book_xlsx.save(output_path)

def json_csv(input_path, output_path):
    # set path to file
    p = Path(r'{}'.format(input_path))
    # read json
    with p.open('r', encoding='utf-8') as f:
        data = json.loads(f.read())
    df = pd.json_normalize(data)
    df.to_csv(r'{}'.format(output_path), index=False, encoding='utf-8')

def json_xlsx(input_path, output_pathpa):
    df_json = pd.read_json(r'{}'.format(input_path))
    df_json.to_excel(r'{}'.format(output_path))

